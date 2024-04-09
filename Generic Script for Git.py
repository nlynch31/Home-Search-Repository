import re
import os
import time
from openpyxl import Workbook, load_workbook
from psutil import process_iter
from openpyxl.utils.exceptions import InvalidFileException

processed_files = []

def extract_info(text):
    # Extracting price
    price_match = re.search(r'\$(\d+(?:,\d+)?)', text)
    price = price_match.group(1) if price_match else None

    # Extracting full address
    address_match = re.search(r'\n(.+?Brooklyn, NY .+)', text)
    full_address = address_match.group(1) if address_match else None

    # Extracting bedrooms and bathrooms
    bedrooms_match = re.search(r'(\d+)\s+beds?', text)
    bedrooms = bedrooms_match.group(1) if bedrooms_match else None
    bathrooms_match = re.search(r'(\d+)\s+baths?', text)
    bathrooms = bathrooms_match.group(1) if bathrooms_match else None

    # Extracting HOA fees
    hoa_match = re.search(r'\$([\d,]+)/mo\sHOA', text)
    hoa_fees = hoa_match.group(1) if hoa_match else None

    # Extracting realtor name
    realtor_match = re.search(r'Listing by:\s*(.+?)\s*Licensed', text, re.DOTALL)
    realtor_name = realtor_match.group(1).strip() if realtor_match else None

    # Updated regular expression for market value
    market_value_match = re.search(r'\$([0-9,]{6,7})\s*Zestimate', text, re.IGNORECASE)
    market_value = market_value_match.group(1) if market_value_match else "N/A"

    # Updated regular expression for description
    description_match = re.search(r'What\'s\s+special\s*[\n\r]+((?:(?!Hide).)*)', text, re.IGNORECASE | re.DOTALL)
    description = description_match.group(1).strip() if description_match else "N/A"

    # Extracting year built
    year_built_match = re.search(r'Built in (\d{4})', text)
    year_built = year_built_match.group(1) if year_built_match else None

    # Extracting subdivision
    subdivision_match = re.search(r'Subdivision: (.+)', text)
    subdivision = subdivision_match.group(1) if subdivision_match else None

    # Extracting URL
    url_match = re.search(r'(https?://\S+)', text)
    url = url_match.group(1) if url_match else None

    # Extracting square feet
    square_feet_match = re.search(r'([\d,]+)\ssqft', text)
    square_feet = square_feet_match.group(1) if square_feet_match else None

    return (
        price, full_address, bedrooms, bathrooms, hoa_fees,
        realtor_name, market_value,
        description, year_built, subdivision, url, square_feet
    )

def process_text_file(file_path):
    """Process a text file and extract information."""
    with open(file_path, 'r', encoding='utf-8') as file:
        text = file.read()
        return extract_info(text)

def is_file_open(file_path):
    """Check if a file is open."""
    try:
        with open(file_path, 'r+', encoding='utf-8') as file:
            return False
    except IOError:
        return True

def save_processed_files():
    """Save processed files list to a file."""
    with open('processed_files.txt', 'w') as file:
        file.write('\n'.join(processed_files))

def load_processed_files():
    """Load processed files list from a file."""
    if os.path.exists('processed_files.txt'):
        with open('processed_files.txt', 'r') as file:
            return file.read().splitlines()
    else:
        return []

def main():
    global processed_files

    # Load processed files list
    processed_files = load_processed_files()

    # Create or load an existing Excel workbook and select the active worksheet
    excel_file_path = 'property_info.xlsx'
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
    else:
        wb = Workbook()
        ws = wb.active

        # Define headers
        headers = ['File Name', 'Price', 'Address', 'Bedrooms', 'Bathrooms', 'HOA Fees',
                   'Realtor Name', 'Market Value', 'Description', 'Year Built', 'Subdivision', 'URL', "Square Feet"]

        # Write headers to the first row
        ws.append(headers)

        # Save the workbook
        wb.save(excel_file_path)

    ws = wb.active

    while True:
        # Check if Excel file is open
        while is_file_open(excel_file_path):
            print("Excel file is open. Pausing program until it is closed.")
            time.sleep(60)  # Check every minute if the file is closed

        # Process all text files in a directory
        directory = r'Your_Filepath_to_Text_Files'
        print("Checking for new text files...")
        for filename in os.listdir(directory):
            if filename.endswith('.txt') and filename not in processed_files:
                print(f"Found new file: {filename}")
                file_path = os.path.join(directory, filename)
                # Process each new text file and append information to the Excel sheet
                info = [filename] + list(process_text_file(file_path))
                ws.append(info)
                processed_files.append(filename)

        try:
            # Save the workbook
            wb.save(excel_file_path)
            # Save processed files list
            save_processed_files()
        except Exception as e:
            print("Error saving Excel file:", e)

        # Sleep for a while before checking again
        time.sleep(60)  # Check every minute

if __name__ == "__main__":
    main()